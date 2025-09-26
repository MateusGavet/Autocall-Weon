# -*- coding: utf-8 -*-

import tkinter as tk
from tkinter import simpledialog, messagebox, scrolledtext
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
import re
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.keys import Keys
import threading
import logging
import os
import sys
import traceback
import time
from datetime import datetime, timedelta
import pyperclip

# --- CONFIGURAÇÕES E CONSTANTES ---
logging.basicConfig(filename="log_automacao.log", level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

# Arquivos de configuração
NOME_ARQUIVO_EXCEL = "automacao_weon.xlsx"
NOME_ARQUIVO_LOGIN = "login.txt"

# Abas da Planilha
ABA_CONTATOS = "contatos"
ABA_RESULTADOS = "resultados"
ABA_RETORNOS = "retornos"
ABA_PRIORIDADE = "PRIORIDADE"

# Seletor do Selenium
SELETOR_BOTAO_ENCERRAR_CHAMADA = "//button[@class='v-btn v-btn--block theme--dark green accent-5']"


# --- FUNÇÃO DE AJUDA PARA PYINSTALLER ---
def resource_path(relative_path):
    """ Retorna o caminho absoluto para o recurso, funciona para desenvolvimento e para o PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# --- FUNÇÕES AUXILIARES ---
def verificar_ou_criar_planilha():
    if not os.path.exists(NOME_ARQUIVO_EXCEL):
        logging.info(f"Arquivo '{NOME_ARQUIVO_EXCEL}' não encontrado. Criando um novo.")
        wb = Workbook()
        ws_contatos = wb.active
        ws_contatos.title = ABA_CONTATOS
        ws_contatos.append(['COD', 'TELEFONE'])
        ws_resultados = wb.create_sheet(ABA_RESULTADOS)
        ws_resultados.append(['COD', 'TELEFONE', 'HORA', 'DATA', 'OBSERVACAO'])
        ws_retornos = wb.create_sheet(ABA_RETORNOS)
        ws_retornos.append(['COD', 'TELEFONE', 'HORA', 'DATA', 'STATUS'])
        ws_prioridade = wb.create_sheet(ABA_PRIORIDADE)
        ws_prioridade.append(['COD'])
        wb.save(NOME_ARQUIVO_EXCEL)
        logging.info("Planilha criada com sucesso.")
        messagebox.showinfo("Planilha Criada", 
                            f"A planilha '{NOME_ARQUIVO_EXCEL}' foi criada com as abas necessárias, incluindo a aba 'PRIORIDADE'.")
    else:
        logging.info(f"Planilha '{NOME_ARQUIVO_EXCEL}' encontrada.")

def verificar_ou_criar_login():
    if not os.path.exists(NOME_ARQUIVO_LOGIN):
        with open(NOME_ARQUIVO_LOGIN, "w", encoding="utf-8") as f:
            f.write("Usuário=\nSenha=\nURL=\n")
        logging.info("Arquivo login.txt criado automaticamente.")
        messagebox.showinfo("Login Criado", f"O arquivo '{NOME_ARQUIVO_LOGIN}' foi criado.\n\nPor favor, preencha seu usuário, senha e a URL do sistema Weon antes de continuar.")

def ler_login():
    try:
        config = {}
        with open(NOME_ARQUIVO_LOGIN, "r", encoding="utf-8") as f:
            for linha in f:
                if "=" in linha:
                    chave, valor = linha.strip().split("=", 1)
                    config[chave.strip().lower()] = valor.strip()
        
        usuario = config.get('usuário')
        senha = config.get('senha')
        url = config.get('url')

        if not all([usuario, senha, url]):
            raise ValueError("Usuário, senha ou URL não preenchidos no login.txt")
        
        return usuario, senha, url
    except (FileNotFoundError, ValueError) as e:
        raise ValueError(f"Erro ao ler '{NOME_ARQUIVO_LOGIN}': {e}.")

def esperar_elemento(driver, by, value, tempo=10):
    try:
        return WebDriverWait(driver, tempo).until(EC.presence_of_element_located((by, value)))
    except TimeoutException:
        logging.error(f"Timeout esperando elemento: {value}")
        return None

def esperar_elemento_clickable(driver, by, value, tempo=10):
    try:
        return WebDriverWait(driver, tempo).until(EC.element_to_be_clickable((by, value)))
    except TimeoutException:
        logging.error(f"Timeout esperando elemento clicável: {value}")
        return None

def limpar_codigo(codigo_bruto):
    if not isinstance(codigo_bruto, str): return ""
    return re.sub(r'[^0-9]', '', codigo_bruto)


# --- CLASSES DE DIÁLOGO ---
class AgendamentoDialog(simpledialog.Dialog):
    def __init__(self, parent, title=None):
        self.result = None
        super().__init__(parent, title)

    def body(self, master):
        self.geometry("320x160")
        tk.Label(master, text="Selecione a data do retorno:", font=("Helvetica", 10, "bold")).grid(row=0, columnspan=3, pady=5)
        self.selected_date = tk.StringVar()
        tk.Button(master, text="Hoje", command=lambda: self.set_date(0)).grid(row=1, column=0, padx=10, pady=5, sticky="ew")
        tk.Button(master, text="Amanhã", command=lambda: self.set_date(1)).grid(row=1, column=1, padx=10, pady=5, sticky="ew")
        tk.Button(master, text="Depois de Amanhã", command=lambda: self.set_date(2)).grid(row=1, column=2, padx=10, pady=5, sticky="ew")
        self.date_label = tk.Label(master, textvariable=self.selected_date, fg="blue", font=("Helvetica", 10, "italic"))
        self.date_label.grid(row=2, columnspan=3, pady=5)
        tk.Label(master, text="Hora (HH:MM):", font=("Helvetica", 10, "bold")).grid(row=3, column=0, pady=10, sticky="e")
        self.time_entry = tk.Entry(master, width=15)
        self.time_entry.grid(row=3, column=1, columnspan=2, pady=10, sticky="w")
        return self.time_entry

    def set_date(self, days_offset):
        target_date = datetime.now() + timedelta(days=days_offset)
        self.selected_date.set(target_date.strftime("%d/%m/%Y"))

    def validate(self):
        if not self.selected_date.get():
            messagebox.showwarning("Erro", "Por favor, selecione uma data.", parent=self)
            return 0
        try:
            time.strptime(self.time_entry.get(), '%H:%M')
            return 1
        except ValueError:
            messagebox.showwarning("Erro de Formato", "Formato de hora inválido. Use HH:MM.", parent=self)
            return 0

    def apply(self):
        self.result = (self.selected_date.get(), self.time_entry.get())

class AddCodDialog(simpledialog.Dialog):
    def body(self, master):
        self.geometry("350x300")
        tk.Label(master, text="Cole os CODs abaixo (um por linha):").pack(pady=10)
        self.text_widget = scrolledtext.ScrolledText(master, width=40, height=10, wrap=tk.WORD)
        self.text_widget.pack(pady=5, padx=10, fill="both", expand=True)
        return self.text_widget

    def apply(self):
        text_content = self.text_widget.get("1.0", tk.END)
        cods = [line.strip() for line in text_content.splitlines() if line.strip()]
        self.result = cods if cods else None


# --- CLASSE PRINCIPAL DA GUI E AUTOMAÇÃO ---
class AutomacaoGUI:
    def __init__(self, master):
        self.master = master
        master.title(f"Automação Weon Integrada by Gavet © {datetime.now().year}")
        master.geometry("550x380")
        master.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        self.is_running = False
        self.is_paused = False
        self.automation_thread = None
        self.driver = None
        self.current_cod = None
        self.current_phone = None
        
        self.priority_queue = []
        self.list_lock = threading.Lock()

        self.excel_lock = threading.Lock()
        self.driver_lock = threading.Lock()
        self.action_taken_event = threading.Event()

        self.status_var = tk.StringVar(value="Status: Ocioso")
        self.cod_var = tk.StringVar(value="COD: -")
        self.fone_var = tk.StringVar(value="Fone: -")
        self.contador_var = tk.StringVar(value="Contatos: 0/0")
        
        tk.Label(master, textvariable=self.status_var, font=("Helvetica", 12, "bold")).pack(pady=10)
        info_frame = tk.Frame(master)
        info_frame.pack(pady=5)
        tk.Label(info_frame, textvariable=self.cod_var, font=("Helvetica", 10)).pack(side=tk.LEFT, padx=10)
        tk.Label(info_frame, textvariable=self.fone_var, font=("Helvetica", 10)).pack(side=tk.LEFT, padx=10)
        tk.Label(master, textvariable=self.contador_var, font=("Helvetica", 10, "italic")).pack(pady=5)

        control_frame = tk.Frame(master)
        control_frame.pack(pady=10)
        self.start_button = tk.Button(control_frame, text="Iniciar", command=self.iniciar_automacao)
        self.start_button.pack(side=tk.LEFT, padx=10)
        self.pause_button = tk.Button(control_frame, text="Pausar", command=self.pausar_automacao, state=tk.DISABLED)
        self.pause_button.pack(side=tk.LEFT, padx=10)
        
        self.add_cod_button = tk.Button(control_frame, text="Adicionar CODs", command=self.adicionar_novos_cods)
        self.add_cod_button.pack(side=tk.LEFT, padx=10)
        
        action_frame = tk.Frame(master)
        action_frame.pack(pady=20)
        self.end_call_button = tk.Button(action_frame, text="Registrar Obs. e Próximo", command=self.registrar_observacao, state=tk.DISABLED)
        self.end_call_button.pack(side=tk.LEFT, padx=5)
        self.schedule_button = tk.Button(action_frame, text="Agendar Retorno", command=self.agendar_retorno, state=tk.DISABLED)
        self.schedule_button.pack(side=tk.LEFT, padx=5)
        
        self.copy_button = tk.Button(master, text="Copiar Código", command=self.copiar_codigo_atual)
        self.copy_button.pack(pady=10)

    def iniciar_automacao(self):
        if self.is_paused:
            self.is_paused = False
            self.start_button.config(state=tk.DISABLED)
            self.pause_button.config(text="Pausar", state=tk.NORMAL)
            self.atualizar_status("Continuando automação...")
        elif not self.is_running:
            self.is_running = True
            self.start_button.config(text="Continuar", state=tk.DISABLED)
            self.pause_button.config(state=tk.NORMAL)
            self.add_cod_button.config(state=tk.NORMAL)
            self.automation_thread = threading.Thread(target=self.loop_principal_automacao, daemon=True)
            self.automation_thread.start()

    def pausar_automacao(self):
        self.is_paused = True
        self.atualizar_status("Pausado")
        self.start_button.config(state=tk.NORMAL)
        self.pause_button.config(text="Continuar", state=tk.DISABLED)

    def on_closing(self):
        if messagebox.askokcancel("Sair", "Deseja fechar a automação?"):
            self.is_running = False
            self.action_taken_event.set() 
            if self.driver:
                try: self.driver.quit()
                except Exception as e: logging.warning(f"Erro ao fechar o Chrome: {e}")
            self.master.destroy()

    def atualizar_status(self, mensagem):
        self.master.after(0, self.status_var.set, f"Status: {mensagem}")
        logging.info(mensagem)
    
    def copiar_codigo_atual(self):
        if self.current_cod:
            pyperclip.copy(self.current_cod)
            self.atualizar_status(f"Código {self.current_cod} copiado.")
        else:
            self.atualizar_status("Nenhum código para copiar.")

    def adicionar_novos_cods(self):
        dialog = AddCodDialog(self.master, "Adicionar CODs com Prioridade")
        novos_cods = dialog.result
        
        if novos_cods:
            self.atualizar_status(f"Adicionando {len(novos_cods)} CODs à fila...")
            
            novas_tarefas = []
            for cod in novos_cods:
                self._escrever_em_planilha(ABA_CONTATOS, [cod, ''])
                novas_tarefas.append({'COD': cod, 'TELEFONE': ''})

            with self.list_lock:
                self.priority_queue.extend(novas_tarefas)
            
            self.atualizar_status(f"{len(novos_cods)} CODs adicionados com prioridade.")

    def registrar_observacao(self):
        obs = simpledialog.askstring("Observação", "Digite a observação da ligação:", parent=self.master)
        self.escrever_resultado(self.current_cod, self.current_phone, obs or "N/A")
        self.action_taken_event.set()

    def agendar_retorno(self):
        dialog = AgendamentoDialog(self.master, "Agendar Retorno")
        if dialog.result:
            data, hora = dialog.result
            nova_linha = [self.current_cod, self.current_phone, hora, data, 'aguardando']
            self._escrever_em_planilha(ABA_RETORNOS, nova_linha)
            obs_agendamento = f"RETORNO AGENDADO para {data} às {hora}"
            self.escrever_resultado(self.current_cod, self.current_phone, obs_agendamento)
            self.action_taken_event.set()

    def escrever_resultado(self, cod, tel, obs):
        agora = datetime.now()
        linha = [str(cod), str(tel), agora.strftime("%H:%M:%S"), agora.strftime("%d/%m/%Y"), obs]
        self._escrever_em_planilha(ABA_RESULTADOS, linha)

    def _escrever_em_planilha(self, nome_aba, dados_linha):
        with self.excel_lock:
            try:
                wb = load_workbook(NOME_ARQUIVO_EXCEL)
                sheet = wb[nome_aba]
                sheet.append(dados_linha)
                wb.save(NOME_ARQUIVO_EXCEL)
                logging.info(f"Linha adicionada à aba '{nome_aba}' com sucesso.")
            except Exception as e:
                logging.error(f"Erro ao escrever na planilha '{nome_aba}': {e}")
                self.atualizar_status(f"Erro ao salvar na planilha!")

    def _atualizar_telefone_na_planilha(self, telefone, cod_alvo):
        with self.excel_lock:
            try:
                wb = load_workbook(NOME_ARQUIVO_EXCEL)
                sheet = wb[ABA_CONTATOS]
                for row in range(2, sheet.max_row + 1):
                    if str(sheet.cell(row, 1).value) == str(cod_alvo):
                        sheet.cell(row, 2, value=telefone)
                        break
                wb.save(NOME_ARQUIVO_EXCEL)
                logging.info(f"Telefone '{telefone}' atualizado para o COD {cod_alvo}.")
            except Exception as e:
                logging.error(f"Erro ao ATUALIZAR telefone na planilha: {e}")
    
    def _limpar_aba_excel(self, nome_aba):
        with self.excel_lock:
            try:
                wb = load_workbook(NOME_ARQUIVO_EXCEL)
                if nome_aba in wb.sheetnames:
                    sheet = wb[nome_aba]
                    if sheet.max_row > 1:
                        sheet.delete_rows(2, sheet.max_row)
                    wb.save(NOME_ARQUIVO_EXCEL)
                    logging.info(f"Aba '{nome_aba}' limpa com sucesso.")
            except Exception as e:
                logging.error(f"Erro ao limpar a aba '{nome_aba}': {e}")

    # --- LÓGICA PRINCIPAL DA AUTOMAÇÃO ---
    def setup_automacao(self):
        try:
            self.atualizar_status("Lendo configurações...")
            usuario, senha, url_weon = ler_login()
            
            self.atualizar_status("Abrindo o Chrome...")
            chrome_options = webdriver.ChromeOptions()
            prefs = {"profile.default_content_setting_values.notifications": 2, "profile.default_content_setting_values.media_stream_mic": 1}
            chrome_options.add_experimental_option("prefs", prefs)
            chrome_options.add_argument("--force-device-scale-factor=0.7")
            
            caminho_chromedriver = resource_path("chromedriver.exe")
            service = ChromeService(executable_path=caminho_chromedriver)
            
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            
            self.driver.get(url_weon)
            
            esperar_elemento(self.driver, By.NAME, "login").send_keys(usuario)
            esperar_elemento(self.driver, By.ID, "password").send_keys(senha)
            esperar_elemento_clickable(self.driver, By.XPATH, "//button[contains(., 'Acessar')]").click()
            
            self.atualizar_status("Login efetuado! Aguardando página carregar...")
            esperar_elemento(self.driver, By.XPATH, '//input[@aria-label="Buscar contato"]', 30)
            return True
        except Exception as e:
            self.atualizar_status("Erro no setup inicial!")
            logging.critical("Erro crítico no setup: %s", traceback.format_exc())
            messagebox.showerror("Erro Crítico no Setup", f"Ocorreu um erro: {e}")
            self.is_running = False
            return False

    def buscar_contato_web(self, codigo_limpo_buscado):
        with self.driver_lock:
            try:
                self.atualizar_status(f"Buscando COD: {codigo_limpo_buscado}")
                wait = WebDriverWait(self.driver, 15)
                
                campo_pesquisa = wait.until(EC.presence_of_element_located((By.XPATH, '//input[@aria-label="Buscar contato"]')))
                campo_pesquisa.clear()
                campo_pesquisa.send_keys(codigo_limpo_buscado)
                campo_pesquisa.send_keys(Keys.ENTER)
                
                seletor_cod = (By.XPATH, "//div[contains(@class, 'v-dialog--active')]//tbody/tr/td[1]")
                seletor_telefone = (By.XPATH, "//div[contains(@class, 'v-dialog--active')]//tbody/tr/td[3]")
                
                cod_elemento = wait.until(EC.text_to_be_present_in_element(seletor_cod, codigo_limpo_buscado))
                
                if not cod_elemento:
                    logging.warning(f"Divergência ou timeout para COD {codigo_limpo_buscado}.")
                    cod_encontrado_na_tela = self.driver.find_element(*seletor_cod).text
                    return cod_encontrado_na_tela, None

                telefone_elemento = self.driver.find_element(*seletor_telefone)
                telefone_encontrado = telefone_elemento.text.strip()
                cod_encontrado = self.driver.find_element(*seletor_cod).text.strip()
                
                if telefone_encontrado and telefone_encontrado != "-":
                    webdriver.ActionChains(self.driver).send_keys(Keys.ESCAPE).perform()
                    time.sleep(0.5)
                    return cod_encontrado, telefone_encontrado
                else:
                    webdriver.ActionChains(self.driver).send_keys(Keys.ESCAPE).perform()
                    time.sleep(0.5)
                    return cod_encontrado, None

            except TimeoutException:
                logging.warning(f"Timeout buscando telefone para {codigo_limpo_buscado}. Nenhum resultado encontrado.")
                try: 
                    webdriver.ActionChains(self.driver).send_keys(Keys.ESCAPE).perform()
                except: pass
                return None, None
            except Exception as e:
                logging.error(f"Erro inesperado ao buscar {codigo_limpo_buscado}. Detalhes: {e}")
                return None, None

    def realizar_chamada(self, telefone):
        with self.driver_lock:
            try:
                time.sleep(1)
                self.atualizar_status(f"Discando para {telefone}...")
                botao_discador = esperar_elemento_clickable(self.driver, By.XPATH, "//button[.//i[contains(@class, 'mdi-rocket-launch')]]", 30)
                if not botao_discador: raise Exception("Botão de discador não encontrado")
                botao_discador.click()
                campo_telefone = esperar_elemento(self.driver, By.XPATH, "//input[@type='text' and @placeholder='Telefone']")
                if not campo_telefone: raise Exception("Campo de telefone não encontrado")
                campo_telefone.clear()
                campo_telefone.send_keys(str(telefone))
                botao_ligar = esperar_elemento_clickable(self.driver, By.XPATH, "//span[contains(@class, 'white--text') and text()='Acionar']")
                if not botao_ligar: raise Exception("Botão 'Acionar' não encontrado")
                botao_ligar.click()
                esperar_elemento(self.driver, By.XPATH, SELETOR_BOTAO_ENCERRAR_CHAMADA, 30)
                return True
            except Exception as e:
                logging.error(f"Falha ao tentar discar para {telefone}: {e}")
                return False

    def loop_principal_automacao(self):
        if not self.setup_automacao():
            self.on_closing()
            return

        try:
            self.atualizar_status("Verificando contatos já realizados...")
            try:
                df_resultados = pd.read_excel(NOME_ARQUIVO_EXCEL, sheet_name=ABA_RESULTADOS)
                contatados_anteriormente = set(df_resultados['COD'].astype(str).unique())
                self.atualizar_status(f"{len(contatados_anteriormente)} contatos já estão nos resultados.")
            except Exception as e:
                logging.warning(f"Não foi possível ler a aba de resultados: {e}")
                contatados_anteriormente = set()
            
            try:
                self.atualizar_status("Lendo a aba de Prioridade...")
                df_prioridade = pd.read_excel(NOME_ARQUIVO_EXCEL, sheet_name=ABA_PRIORIDADE, dtype=str)
                df_prioridade.dropna(subset=['COD'], inplace=True)
                cods_prioritarios = df_prioridade['COD'].tolist()
                
                if cods_prioritarios:
                    tarefas_prioritarias = [{'COD': cod, 'TELEFONE': ''} for cod in cods_prioritarios]
                    with self.list_lock:
                        self.priority_queue.extend(tarefas_prioritarias)
                    self.atualizar_status(f"{len(cods_prioritarios)} CODs carregados da aba Prioridade.")
                    self._limpar_aba_excel(ABA_PRIORIDADE)
            except Exception as e:
                logging.warning(f"Não foi possível ler a aba de Prioridade: {e}")

            contatos_df = pd.read_excel(NOME_ARQUIVO_EXCEL, sheet_name=ABA_CONTATOS, dtype=str).fillna('')
            lista_de_tarefas = [row.to_dict() for index, row in contatos_df.iterrows()]
            
            indice_atual = 0
            while self.is_running and (indice_atual < len(lista_de_tarefas) or self.priority_queue):
                while self.is_paused: time.sleep(0.5)
                if not self.is_running: break

                tarefa_atual = None
                
                with self.list_lock:
                    if self.priority_queue:
                        tarefa_atual = self.priority_queue.pop(0)

                if not tarefa_atual:
                    tarefa_atual = lista_de_tarefas[indice_atual]
                    indice_atual += 1

                cod_original = str(tarefa_atual.get('COD', ''))
                telefone_existente = str(tarefa_atual.get('TELEFONE', ''))
                
                total_na_fila = len(lista_de_tarefas) + len(self.priority_queue)
                self.contador_var.set(f"Contatos: {indice_atual}/{total_na_fila}")
                self.master.update_idletasks()

                if not cod_original: continue

                if cod_original in contatados_anteriormente:
                    self.atualizar_status(f"COD {cod_original} já contatado. Pulando.")
                    time.sleep(1)
                    continue

                telefone_para_ligar = telefone_existente

                if not telefone_para_ligar:
                    codigo_limpo = limpar_codigo(cod_original)
                    if len(codigo_limpo) > 5:
                        cod_encontrado, telefone_encontrado = self.buscar_contato_web(codigo_limpo)
                        
                        if cod_encontrado and (limpar_codigo(cod_encontrado) != codigo_limpo):
                            obs = f"DIVERGÊNCIA: Buscou {codigo_limpo}, encontrou {cod_encontrado}"
                            self.escrever_resultado(cod_original, '', obs)
                            contatados_anteriormente.add(cod_original)
                            self.atualizar_status(obs + ". Pulando.")
                            time.sleep(2)
                            continue

                        if telefone_encontrado:
                            self.atualizar_status(f"Telefone encontrado: {telefone_encontrado}")
                            self._atualizar_telefone_na_planilha(telefone_encontrado, cod_original)
                            telefone_para_ligar = telefone_encontrado
                        else:
                            self.escrever_resultado(cod_original, '', 'TELEFONE NÃO ENCONTRADO')
                            contatados_anteriormente.add(cod_original)
                            self.atualizar_status(f"Telefone não encontrado para {cod_original}. Pulando.")
                            time.sleep(2)
                            continue
                    else:
                        self.escrever_resultado(cod_original, '', 'CÓDIGO/CNPJ INVÁLIDO')
                        contatados_anteriormente.add(cod_original)
                        continue
                
                self.current_cod = cod_original
                self.current_phone = telefone_para_ligar
                self.cod_var.set(f"COD: {self.current_cod}")
                self.fone_var.set(f"Fone: {self.current_phone or 'Não encontrado'}")
                self.master.update_idletasks()
                
                if telefone_para_ligar:
                    if self.realizar_chamada(telefone_para_ligar):
                        
                        self.atualizar_status("Em chamada... Aguardando sua ação.")
                        self.master.after(0, self.end_call_button.config, {'state': tk.NORMAL})
                        self.master.after(0, self.schedule_button.config, {'state': tk.NORMAL})
                        
                        self.action_taken_event.clear()
                        self.action_taken_event.wait()
                        
                        contatados_anteriormente.add(cod_original)
                        
                        self.master.after(0, self.end_call_button.config, {'state': tk.DISABLED})
                        self.master.after(0, self.schedule_button.config, {'state': tk.DISABLED})
                    else:
                        self.atualizar_status("Erro ao discar. Próximo em 5s.")
                        self.escrever_resultado(cod_original, telefone_para_ligar, "ERRO AO DISCAR")
                        contatados_anteriormente.add(cod_original)
                        time.sleep(5)
                else:
                    self.atualizar_status(f"Nenhum telefone para {cod_original}. Pulando.")
                    time.sleep(2)

        except Exception as e:
            logging.error(f"Erro no loop principal: {traceback.format_exc()}")
            messagebox.showerror("Erro no Loop", f"Ocorreu um erro inesperado: {e}")
        finally:
            self.atualizar_status("Automação finalizada.")
            self.is_running = False
            self.is_paused = False
            self.start_button.config(text="Iniciar", state=tk.NORMAL)
            self.pause_button.config(text="Pausar", state=tk.DISABLED)
            self.add_cod_button.config(state=tk.NORMAL)

# --- PONTO DE ENTRADA DO PROGRAMA ---
if __name__ == "__main__":
    verificar_ou_criar_login()
    verificar_ou_criar_planilha()
    root = tk.Tk()
    app = AutomacaoGUI(root)
    root.mainloop()